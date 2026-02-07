"""
PolicyHub Users Settings View (PySide6)

User management view for administrators to add, edit, and manage users.
"""

from pathlib import Path
from typing import TYPE_CHECKING, List

from PySide6.QtCore import Qt, QTimer
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QLineEdit,
    QFrame,
    QCheckBox,
    QTableView,
    QHeaderView,
    QAbstractItemView,
    QMenu,
    QFileDialog,
)
from PySide6.QtGui import QColor

from app.theme import COLORS, SPACING, TYPOGRAPHY, style_button, style_card
from ui.components.empty_state import EmptyState
from ui.components.toast import Toast
from ui.dialogs.confirm_dialog import ConfirmDialog, InfoDialog
from ui.models.user_table_model import UserTableModel
from ui.delegates.status_delegate import RoleBadgeDelegate, ActiveStatusDelegate
from ui.views.base_view import BaseView
from models.user import User
from services.user_service import UserService

if TYPE_CHECKING:
    from app.application import PolicyHubApp

# Try to import openpyxl for export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class UsersSettingsView(BaseView):
    """
    User management view for administrators.

    Features:
    - View all users in a table
    - Add new users
    - Edit existing users
    - Activate/deactivate users
    - Reset passwords
    - Import/export users
    """

    def __init__(self, parent: QWidget, app: "PolicyHubApp"):
        """
        Initialize the users settings view.

        Args:
            parent: Parent widget
            app: Main application instance
        """
        super().__init__(parent, app)
        self.user_service = UserService(app.db)
        self.users: List[User] = []
        self.filtered_users: List[User] = []
        self.show_inactive = False
        self.search_term = ""
        self._search_timer = QTimer()
        self._search_timer.setSingleShot(True)
        self._search_timer.timeout.connect(self._perform_search)
        self._build_ui()

    def _build_ui(self) -> None:
        """Build the users management UI."""
        layout = QVBoxLayout(self)
        layout.setContentsMargins(
            SPACING.WINDOW_PADDING,
            SPACING.WINDOW_PADDING,
            SPACING.WINDOW_PADDING,
            SPACING.WINDOW_PADDING,
        )
        layout.setSpacing(10)

        # Header
        header = QFrame()
        header.setStyleSheet("background: transparent;")
        header_layout = QHBoxLayout(header)
        header_layout.setContentsMargins(0, 0, 0, 0)

        title = QLabel("User Management")
        title.setFont(TYPOGRAPHY.section_heading)
        title.setStyleSheet(f"color: {COLORS.TEXT_PRIMARY}; background: transparent;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        # Add button with icon
        add_btn = QPushButton("➕ Add User")
        add_btn.setFixedSize(120, 36)
        style_button(add_btn, "primary")
        add_btn.clicked.connect(self._on_add_user)
        header_layout.addWidget(add_btn)

        layout.addWidget(header)

        # Toolbar
        toolbar = QFrame()
        toolbar.setStyleSheet("background: transparent;")
        toolbar_layout = QHBoxLayout(toolbar)
        toolbar_layout.setContentsMargins(0, 0, 0, 0)
        toolbar_layout.setSpacing(8)

        # Search with icon in placeholder
        self.search_entry = QLineEdit()
        self.search_entry.setFixedSize(250, 32)
        self.search_entry.setPlaceholderText("🔍 Search users...")
        self.search_entry.setFont(TYPOGRAPHY.body)
        self._style_input(self.search_entry)
        self.search_entry.textChanged.connect(self._on_search_changed)
        toolbar_layout.addWidget(self.search_entry)

        # Show inactive checkbox
        self.inactive_checkbox = QCheckBox("Show inactive")
        self.inactive_checkbox.setFont(TYPOGRAPHY.small)
        self.inactive_checkbox.setStyleSheet(f"color: {COLORS.TEXT_SECONDARY}; background: transparent;")
        self.inactive_checkbox.stateChanged.connect(self._refresh_table)
        toolbar_layout.addWidget(self.inactive_checkbox)

        toolbar_layout.addStretch()

        # Export button with icon
        export_btn = QPushButton("📥 Export")
        export_btn.setFixedSize(90, 32)
        style_button(export_btn, "secondary")
        export_btn.clicked.connect(self._on_export_users)
        toolbar_layout.addWidget(export_btn)

        # Template button with icon
        template_btn = QPushButton("📋 Template")
        template_btn.setFixedSize(100, 32)
        style_button(template_btn, "secondary")
        template_btn.clicked.connect(self._on_download_template)
        toolbar_layout.addWidget(template_btn)

        # Import button with icon
        import_btn = QPushButton("📤 Import")
        import_btn.setFixedSize(90, 32)
        style_button(import_btn, "secondary")
        import_btn.clicked.connect(self._on_import_users)
        toolbar_layout.addWidget(import_btn)

        # Visual separator
        separator = QFrame()
        separator.setProperty("toolbarSeparator", True)
        separator.setFixedHeight(24)
        separator.setStyleSheet(f"""
            QFrame {{
                background-color: {COLORS.BORDER};
                max-width: 1px;
                min-width: 1px;
            }}
        """)
        toolbar_layout.addWidget(separator)

        # Bulk deactivate button with icon
        self.bulk_deactivate_btn = QPushButton("🚫 Deactivate")
        self.bulk_deactivate_btn.setFixedSize(110, 32)
        style_button(self.bulk_deactivate_btn, "secondary")
        self.bulk_deactivate_btn.setEnabled(False)
        self.bulk_deactivate_btn.clicked.connect(self._on_bulk_deactivate)
        toolbar_layout.addWidget(self.bulk_deactivate_btn)

        layout.addWidget(toolbar)

        # Auth Settings Card
        self._build_auth_settings_card(layout)

        # Table
        table_card = QFrame()
        style_card(table_card, with_shadow=True)
        table_layout = QVBoxLayout(table_card)
        table_layout.setContentsMargins(
            SPACING.CARD_PADDING,
            SPACING.CARD_PADDING,
            SPACING.CARD_PADDING,
            SPACING.CARD_PADDING,
        )
        table_layout.setSpacing(0)

        # Table view
        self.table_model = UserTableModel()
        self.table_view = QTableView()
        self.table_view.setModel(self.table_model)

        # Configure table
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table_view.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self.table_view.setShowGrid(True)
        self.table_view.verticalHeader().setVisible(False)
        self.table_view.verticalHeader().setDefaultSectionSize(36)

        # Header
        header_view = self.table_view.horizontalHeader()
        header_view.setStretchLastSection(True)
        header_view.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header_view.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)

        # Column widths
        self.table_view.setColumnWidth(0, 120)
        self.table_view.setColumnWidth(1, 180)
        self.table_view.setColumnWidth(2, 200)
        self.table_view.setColumnWidth(3, 100)
        self.table_view.setColumnWidth(4, 80)
        self.table_view.setColumnWidth(5, 140)

        # Delegates
        role_delegate = RoleBadgeDelegate(self.table_view)
        self.table_view.setItemDelegateForColumn(3, role_delegate)
        status_delegate = ActiveStatusDelegate(self.table_view)
        self.table_view.setItemDelegateForColumn(4, status_delegate)

        # Signals
        self.table_view.doubleClicked.connect(self._on_double_click)
        self.table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self._show_context_menu)
        self.table_view.selectionModel().selectionChanged.connect(self._on_selection_changed)

        # Styling
        self.table_view.setStyleSheet(f"""
            QTableView {{
                background-color: {COLORS.CARD};
                border: none;
                border-radius: 4px;
                gridline-color: {COLORS.BORDER};
            }}
            QTableView::item {{
                padding: 8px;
            }}
            QTableView::item:selected {{
                background-color: {COLORS.PRIMARY};
                color: {COLORS.PRIMARY_FOREGROUND};
            }}
            QHeaderView::section {{
                background-color: {COLORS.MUTED};
                color: {COLORS.TEXT_PRIMARY};
                font-weight: bold;
                padding: 8px;
                border: none;
                border-bottom: 1px solid {COLORS.BORDER};
            }}
        """)

        table_layout.addWidget(self.table_view)

        # Empty state (hidden by default)
        self.empty_state = EmptyState(
            parent=table_card,
            icon=EmptyState.ICON_USER,
            title="No users found",
            message="No users match your search criteria.",
        )
        self.empty_state.hide()
        table_layout.addWidget(self.empty_state)

        # Status bar
        self.status_label = QLabel("Loading...")
        self.status_label.setFont(TYPOGRAPHY.small)
        self.status_label.setStyleSheet(f"color: {COLORS.TEXT_SECONDARY}; background: transparent;")
        table_layout.addWidget(self.status_label)

        layout.addWidget(table_card, 1)

    def _build_auth_settings_card(self, parent_layout: QVBoxLayout) -> None:
        """Build the authentication settings card."""
        from services.settings_service import SettingsService

        self.settings_service = SettingsService(self.app.db)

        card = QFrame()
        style_card(card, with_shadow=True)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(
            SPACING.CARD_PADDING,
            SPACING.CARD_PADDING,
            SPACING.CARD_PADDING,
            SPACING.CARD_PADDING,
        )
        card_layout.setSpacing(12)

        # Title
        title = QLabel("Authentication Settings")
        title.setFont(TYPOGRAPHY.section_heading)
        title.setStyleSheet(f"color: {COLORS.TEXT_PRIMARY}; background: transparent;")
        card_layout.addWidget(title)

        # Settings row
        settings_row = QHBoxLayout()

        # Left: Require Login
        left_frame = QWidget()
        left_frame.setStyleSheet("background: transparent;")
        left_layout = QVBoxLayout(left_frame)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(4)

        self.require_login_checkbox = QCheckBox("Require Login")
        self.require_login_checkbox.setFont(TYPOGRAPHY.body)
        self.require_login_checkbox.setStyleSheet(f"color: {COLORS.TEXT_PRIMARY}; background: transparent;")
        self.require_login_checkbox.setChecked(self.settings_service.get_require_login())
        self.require_login_checkbox.stateChanged.connect(self._on_require_login_changed)
        left_layout.addWidget(self.require_login_checkbox)

        help_text = QLabel("When disabled, the app will automatically log in as the admin user.")
        help_text.setFont(TYPOGRAPHY.small)
        help_text.setStyleSheet(f"color: {COLORS.TEXT_MUTED}; background: transparent;")
        left_layout.addWidget(help_text)

        settings_row.addWidget(left_frame)
        settings_row.addStretch()

        # Right: Master Password
        right_frame = QWidget()
        right_frame.setStyleSheet("background: transparent;")
        right_layout = QVBoxLayout(right_frame)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(4)
        right_layout.setAlignment(Qt.AlignmentFlag.AlignRight)

        master_pwd_btn = QPushButton("Change Master Password")
        master_pwd_btn.setFixedSize(180, 32)
        style_button(master_pwd_btn, "secondary")
        master_pwd_btn.clicked.connect(self._on_change_master_password)
        right_layout.addWidget(master_pwd_btn, 0, Qt.AlignmentFlag.AlignRight)

        pwd_help = QLabel("Safety net for forgotten admin passwords")
        pwd_help.setFont(TYPOGRAPHY.small)
        pwd_help.setStyleSheet(f"color: {COLORS.TEXT_MUTED}; background: transparent;")
        right_layout.addWidget(pwd_help, 0, Qt.AlignmentFlag.AlignRight)

        settings_row.addWidget(right_frame)

        card_layout.addLayout(settings_row)
        parent_layout.addWidget(card)

    def _style_input(self, widget: QLineEdit) -> None:
        """Apply input styling."""
        widget.setStyleSheet(f"""
            QLineEdit {{
                background-color: {COLORS.BACKGROUND};
                border: 1px solid {COLORS.BORDER};
                border-radius: 4px;
                padding: 0 12px;
                color: {COLORS.TEXT_PRIMARY};
            }}
            QLineEdit:focus {{
                border-color: {COLORS.PRIMARY};
            }}
        """)

    def _on_require_login_changed(self, state: int) -> None:
        """Handle require login toggle change."""
        require_login = state == Qt.CheckState.Checked.value

        try:
            self.settings_service.set_require_login(require_login)

            status = "enabled" if require_login else "disabled"
            Toast.show_success(self, f"Login requirement {status}")

        except PermissionError:
            self.require_login_checkbox.setChecked(not require_login)
            Toast.show_error(self, "Permission denied")
        except Exception as e:
            self.require_login_checkbox.setChecked(not require_login)
            Toast.show_error(self, f"Failed to update setting: {str(e)}")

    def _on_change_master_password(self) -> None:
        """Handle change master password button click."""
        from ui.dialogs.master_password_dialog import MasterPasswordDialog

        dialog = MasterPasswordDialog(self.window(), self.app.db)
        dialog.exec()

    def _refresh_table(self) -> None:
        """Reload user data."""
        self.show_inactive = self.inactive_checkbox.isChecked()
        self.users = self.user_service.get_all_users(include_inactive=self.show_inactive)
        self._apply_filter()

    def _apply_filter(self) -> None:
        """Apply search filter to users and update table."""
        search = self.search_term.lower().strip()

        if search:
            self.filtered_users = [
                u for u in self.users
                if search in u.username.lower()
                or search in u.full_name.lower()
                or (u.email and search in u.email.lower())
            ]
        else:
            self.filtered_users = self.users.copy()

        self.table_model.set_users(self.filtered_users)
        self.table_view.clearSelection()

        # Show/hide empty state
        has_users = len(self.filtered_users) > 0
        self.table_view.setVisible(has_users)
        self.empty_state.setVisible(not has_users)

        # Update status label
        if search and not has_users:
            self.status_label.setText("No users match your search")
        else:
            self.status_label.setText(f"Showing {len(self.filtered_users)} user(s)")

        self.bulk_deactivate_btn.setEnabled(False)

    def _on_search_changed(self, text: str) -> None:
        """Handle search input changes with debouncing."""
        self._search_timer.stop()
        self._search_timer.start(300)

    def _perform_search(self) -> None:
        """Perform the actual search."""
        self.search_term = self.search_entry.text()
        self._apply_filter()

    def _on_selection_changed(self) -> None:
        """Handle table selection changes."""
        selected = self.table_view.selectionModel().selectedRows()
        self.bulk_deactivate_btn.setEnabled(len(selected) > 0)

    def _on_add_user(self) -> None:
        """Handle add user button click."""
        from ui.dialogs.user_dialog import UserDialog

        dialog = UserDialog(self.window(), self.app.db)
        if dialog.exec():
            Toast.show_success(self, "User created successfully")
            self._refresh_table()

    def _on_double_click(self, index) -> None:
        """Handle double-click on table row."""
        user = self.table_model.get_user(index.row())
        if user:
            self._edit_user(user)

    def _edit_user(self, user: User) -> None:
        """Open edit dialog for a user."""
        from ui.dialogs.user_dialog import UserDialog

        dialog = UserDialog(self.window(), self.app.db, user=user)
        if dialog.exec():
            self._refresh_table()

    def _show_context_menu(self, pos) -> None:
        """Show context menu on right-click."""
        index = self.table_view.indexAt(pos)
        if not index.isValid():
            return

        user = self.table_model.get_user(index.row())
        if not user:
            return

        menu = QMenu(self)
        menu.setStyleSheet(f"""
            QMenu {{
                background-color: {COLORS.CARD};
                border: 1px solid {COLORS.BORDER};
                border-radius: 4px;
                padding: 4px;
            }}
            QMenu::item {{
                padding: 8px 16px;
                border-radius: 4px;
            }}
            QMenu::item:selected {{
                background-color: {COLORS.MUTED};
            }}
        """)

        edit_action = menu.addAction("Edit User")
        edit_action.triggered.connect(lambda: self._edit_user(user))

        if user.is_active:
            deactivate_action = menu.addAction("Deactivate")
            deactivate_action.triggered.connect(lambda: self._toggle_active(user, False))
        else:
            activate_action = menu.addAction("Activate")
            activate_action.triggered.connect(lambda: self._toggle_active(user, True))

        reset_action = menu.addAction("Reset Password")
        reset_action.triggered.connect(lambda: self._reset_password(user))

        menu.exec(self.table_view.viewport().mapToGlobal(pos))

    def _toggle_active(self, user: User, activate: bool) -> None:
        """Activate or deactivate a user."""
        try:
            if activate:
                self.user_service.activate_user(user.user_id)
                Toast.show_success(self, f"User '{user.username}' activated")
            else:
                # Check if trying to deactivate self
                from core.session import SessionManager
                session = SessionManager.get_instance()
                if session.current_user and session.current_user.user_id == user.user_id:
                    Toast.show_error(self, "You cannot deactivate your own account")
                    return

                self.user_service.deactivate_user(user.user_id)
                Toast.show_success(self, f"User '{user.username}' deactivated")
            self._refresh_table()
        except Exception as e:
            Toast.show_error(self, str(e))

    def _reset_password(self, user: User) -> None:
        """Reset a user's password."""
        from ui.dialogs.password_reset_dialog import PasswordResetDialog

        dialog = PasswordResetDialog(self.window(), self.app.db, user=user)
        if dialog.exec():
            InfoDialog.show_success(
                self.window(),
                "Password Reset",
                f"Password for '{user.username}' has been reset successfully.",
            )

    def _on_export_users(self) -> None:
        """Export user list to Excel."""
        if not OPENPYXL_AVAILABLE:
            InfoDialog.show_error(
                self.window(),
                "Export Unavailable",
                "Excel export requires openpyxl library.",
            )
            return

        export_data = self.user_service.get_users_for_export(
            include_inactive=self.inactive_checkbox.isChecked()
        )

        if not export_data:
            InfoDialog.show_info(
                self.window(),
                "No Data",
                "No users to export.",
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Users",
            "users_export.xlsx",
            "Excel Files (*.xlsx)",
        )

        if not file_path:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Users"

            # Styles
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin", color="E5E7EB"),
                right=Side(style="thin", color="E5E7EB"),
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB"),
            )

            # Headers
            headers = ["Username", "Full Name", "Email", "Role", "Status", "Created", "Last Login"]
            for col, header_text in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header_text)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Data
            for row_idx, user_data in enumerate(export_data, 2):
                ws.cell(row=row_idx, column=1, value=user_data["username"]).border = thin_border
                ws.cell(row=row_idx, column=2, value=user_data["full_name"]).border = thin_border
                ws.cell(row=row_idx, column=3, value=user_data["email"]).border = thin_border
                ws.cell(row=row_idx, column=4, value=user_data["role"]).border = thin_border
                ws.cell(row=row_idx, column=5, value=user_data["status"]).border = thin_border
                ws.cell(row=row_idx, column=6, value=user_data["created_at"]).border = thin_border
                ws.cell(row=row_idx, column=7, value=user_data["last_login"]).border = thin_border

            # Column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 15

            wb.save(file_path)

            InfoDialog.show_success(
                self.window(),
                "Export Complete",
                f"Exported {len(export_data)} users to:\n{file_path}",
            )

        except Exception as e:
            InfoDialog.show_error(
                self.window(),
                "Export Error",
                f"Failed to export users: {str(e)}",
            )

    def _on_download_template(self) -> None:
        """Download Excel import template."""
        if not OPENPYXL_AVAILABLE:
            InfoDialog.show_error(
                self.window(),
                "Feature Unavailable",
                "Excel template requires openpyxl library.",
            )
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Import Template",
            "user_import_template.xlsx",
            "Excel Files (*.xlsx)",
        )

        if not file_path:
            return

        try:
            self.user_service.create_xlsx_template(Path(file_path))

            InfoDialog.show_success(
                self.window(),
                "Template Saved",
                f"Import template saved to:\n{file_path}\n\n"
                "Instructions:\n"
                "1. Open the template in Excel\n"
                "2. Fill in user data (delete the example row)\n"
                "3. Use the Role dropdown to select roles\n"
                "4. Save and use Import to add users",
            )

        except Exception as e:
            InfoDialog.show_error(
                self.window(),
                "Error",
                f"Failed to save template: {str(e)}",
            )

    def _on_import_users(self) -> None:
        """Import users from Excel or CSV file."""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Users",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*.*)",
        )

        if not file_path:
            return

        # Validate first
        try:
            result = self.user_service.validate_import_data(Path(file_path))
        except Exception as e:
            InfoDialog.show_error(
                self.window(),
                "Validation Error",
                f"Failed to validate file: {str(e)}",
            )
            return

        if not result.success:
            error_text = "\n".join(result.error_messages[:20])
            if len(result.errors) > 20:
                error_text += f"\n... and {len(result.errors) - 20} more errors"

            InfoDialog.show_error(
                self.window(),
                "Import Validation Failed",
                f"The following errors were found:\n\n{error_text}",
            )
            return

        if result.imported_count == 0:
            InfoDialog.show_error(
                self.window(),
                "No Users Found",
                "No valid user rows were found in the file.\n\n"
                "Make sure you have filled in user data in the template.",
            )
            return

        # Confirm import
        confirm = ConfirmDialog.ask(
            self.window(),
            "Confirm Import",
            f"Ready to import {result.imported_count} user(s).\n\n"
            "Users will be required to change their password on first login.\n\n"
            "Continue with import?",
        )

        if not confirm:
            return

        # Perform import
        try:
            import_result = self.user_service.import_users(Path(file_path))

            if import_result.success:
                Toast.show_success(self, f"Imported {import_result.imported_count} user(s)")
                self._refresh_table()
            else:
                error_text = "\n".join(import_result.error_messages)
                InfoDialog.show_error(
                    self.window(),
                    "Import Failed",
                    f"Import failed:\n\n{error_text}",
                )

        except PermissionError:
            Toast.show_error(self, "Permission denied")
        except Exception as e:
            Toast.show_error(self, f"Failed to import users: {str(e)}")

    def _on_bulk_deactivate(self) -> None:
        """Deactivate selected users."""
        selected_indexes = self.table_view.selectionModel().selectedRows()
        if not selected_indexes:
            return

        selected_users = [
            self.table_model.get_user(idx.row())
            for idx in selected_indexes
            if self.table_model.get_user(idx.row())
        ]

        if not selected_users:
            return

        # Confirm
        usernames = ", ".join(u.username for u in selected_users[:5])
        if len(selected_users) > 5:
            usernames += f" and {len(selected_users) - 5} more"

        confirm = ConfirmDialog.ask(
            self.window(),
            "Confirm Bulk Deactivation",
            f"Are you sure you want to deactivate {len(selected_users)} user(s)?\n\n"
            f"Users: {usernames}",
        )

        if not confirm:
            return

        user_ids = [u.user_id for u in selected_users]

        try:
            count, errors = self.user_service.bulk_deactivate_users(user_ids)

            if errors:
                error_text = "\n".join(errors[:10])
                if len(errors) > 10:
                    error_text += f"\n... and {len(errors) - 10} more"

                InfoDialog.show_info(
                    self.window(),
                    "Partial Success",
                    f"Deactivated {count} user(s).\n\nSome users could not be deactivated:\n{error_text}",
                )
            else:
                InfoDialog.show_success(
                    self.window(),
                    "Users Deactivated",
                    f"Successfully deactivated {count} user(s).",
                )

            self._refresh_table()

        except PermissionError:
            InfoDialog.show_error(
                self.window(),
                "Permission Denied",
                "You do not have permission to deactivate users.",
            )
        except Exception as e:
            InfoDialog.show_error(
                self.window(),
                "Error",
                f"Failed to deactivate users: {str(e)}",
            )

    def on_show(self) -> None:
        """Called when the view becomes visible."""
        super().on_show()
        self._refresh_table()
