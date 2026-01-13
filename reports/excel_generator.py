"""
PolicyHub Excel Generator

Generates Excel reports using openpyxl.
"""

import logging
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reports.base_report import BaseReportGenerator, ReportData

logger = logging.getLogger(__name__)


# Professional color palette (Excel hex format without #)
COLORS = {
    "primary": "1E3A8A",       # Deep blue
    "secondary": "64748B",     # Slate gray
    "header_bg": "1E3A8A",     # Deep blue
    "header_fg": "FFFFFF",     # White
    "row_alt": "F8FAFC",       # Light gray
    "border": "E2E8F0",        # Border gray
    "summary_bg": "F1F5F9",    # Summary background
}


class ExcelGenerator(BaseReportGenerator):
    """
    Generates Excel reports using openpyxl.

    Creates professional workbooks with:
    - Formatted header with title and metadata
    - Summary statistics section
    - Data table with filters enabled
    - Proper column widths and formatting
    """

    def __init__(self, report_data: ReportData):
        """
        Initialize the Excel generator.

        Args:
            report_data: Report configuration and data
        """
        super().__init__(report_data)
        self.wb = Workbook()
        self.ws = self.wb.active
        self._setup_styles()

    def _setup_styles(self) -> None:
        """Set up reusable cell styles."""
        # Fonts
        self.font_title = Font(name="Calibri", size=16, bold=True, color=COLORS["primary"])
        self.font_subtitle = Font(name="Calibri", size=10, color=COLORS["secondary"])
        self.font_header = Font(name="Calibri", size=10, bold=True, color=COLORS["header_fg"])
        self.font_body = Font(name="Calibri", size=10)
        self.font_summary_label = Font(name="Calibri", size=9, color=COLORS["secondary"])
        self.font_summary_value = Font(name="Calibri", size=11, bold=True)

        # Fills
        self.fill_header = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        self.fill_alt_row = PatternFill(start_color=COLORS["row_alt"], end_color=COLORS["row_alt"], fill_type="solid")
        self.fill_summary = PatternFill(start_color=COLORS["summary_bg"], end_color=COLORS["summary_bg"], fill_type="solid")

        # Borders
        thin_side = Side(style="thin", color=COLORS["border"])
        self.border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        # Alignment
        self.align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def generate(self, output_path: str) -> str:
        """
        Generate the Excel report.

        Args:
            output_path: Path to save the Excel file

        Returns:
            Path to the generated file
        """
        self.ws.title = "Report"

        current_row = 1

        # Title
        self.ws.cell(row=current_row, column=1, value=self.data.config.title)
        self.ws.cell(row=current_row, column=1).font = self.font_title
        current_row += 1

        # Subtitle / Generation info
        gen_info = f"Generated: {self.data.generated_at.strftime('%Y-%m-%d %H:%M')}"
        if self.data.config.subtitle:
            subtitle = f"{self.data.config.subtitle} | {gen_info}"
        else:
            subtitle = gen_info
        self.ws.cell(row=current_row, column=1, value=subtitle)
        self.ws.cell(row=current_row, column=1).font = self.font_subtitle
        current_row += 1

        # Filters applied
        if self.data.filters_applied:
            filter_parts = [f"{k}: {v}" for k, v in self.data.filters_applied.items() if v]
            if filter_parts:
                filter_text = "Filters: " + " | ".join(filter_parts)
                self.ws.cell(row=current_row, column=1, value=filter_text)
                self.ws.cell(row=current_row, column=1).font = self.font_subtitle
                current_row += 1

        current_row += 1  # Blank row

        # Summary section
        if self.data.config.include_summary and self.data.summary:
            current_row = self._build_summary_section(current_row)
            current_row += 1  # Blank row

        # Data table
        if self.data.rows:
            current_row = self._build_data_table(current_row)
        else:
            self.ws.cell(row=current_row, column=1, value="No data to display.")
            current_row += 1

        # Footer
        if self.data.config.include_footer:
            current_row += 1
            self.ws.cell(row=current_row, column=1, value=f"Total records: {self.data.row_count}")
            self.ws.cell(row=current_row, column=1).font = self.font_subtitle

        # Auto-adjust column widths
        self._auto_adjust_columns()

        # Save workbook
        self.wb.save(output_path)
        logger.info(f"Excel report generated: {output_path}")

        return output_path

    def _build_summary_section(self, start_row: int) -> int:
        """
        Build the summary statistics section.

        Args:
            start_row: Row to start at

        Returns:
            Next available row
        """
        col = 1
        for key, value in self.data.summary.items():
            # Label
            label_cell = self.ws.cell(row=start_row, column=col, value=key)
            label_cell.font = self.font_summary_label
            label_cell.fill = self.fill_summary
            label_cell.alignment = self.align_center
            label_cell.border = self.border_all

            # Value
            value_cell = self.ws.cell(row=start_row + 1, column=col, value=value)
            value_cell.font = self.font_summary_value
            value_cell.fill = self.fill_summary
            value_cell.alignment = self.align_center
            value_cell.border = self.border_all

            col += 1

        return start_row + 2

    def _build_data_table(self, start_row: int) -> int:
        """
        Build the main data table.

        Args:
            start_row: Row to start at

        Returns:
            Next available row
        """
        # Header row
        for col_idx, col in enumerate(self.data.columns, start=1):
            cell = self.ws.cell(row=start_row, column=col_idx, value=col.header)
            cell.font = self.font_header
            cell.fill = self.fill_header
            cell.border = self.border_all
            cell.alignment = self.align_center

        start_row += 1

        # Data rows
        for row_idx, row_data in enumerate(self.data.rows):
            for col_idx, col in enumerate(self.data.columns, start=1):
                value = row_data.get(col.key, "")
                formatted = self._format_value(value)

                cell = self.ws.cell(row=start_row, column=col_idx, value=formatted)
                cell.font = self.font_body
                cell.border = self.border_all

                # Apply alignment
                if col.align == "center":
                    cell.alignment = self.align_center
                elif col.align == "right":
                    cell.alignment = self.align_right
                else:
                    cell.alignment = self.align_left

                # Alternating row color
                if row_idx % 2 == 1:
                    cell.fill = self.fill_alt_row

            start_row += 1

        # Enable auto-filter
        if self.data.rows:
            header_row = start_row - len(self.data.rows) - 1
            last_col = get_column_letter(len(self.data.columns))
            filter_range = f"A{header_row}:{last_col}{start_row - 1}"
            self.ws.auto_filter.ref = filter_range

        return start_row

    def _auto_adjust_columns(self) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx, column_cells in enumerate(self.ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            for cell in column_cells:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            # Set width with some padding, but cap at reasonable max
            adjusted_width = min(max_length + 2, 50)
            adjusted_width = max(adjusted_width, 10)  # Minimum width
            self.ws.column_dimensions[col_letter].width = adjusted_width
