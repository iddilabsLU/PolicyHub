"""
PolicyHub Users Settings View

User management view for administrators to add, edit, and manage users.
"""

from pathlib import Path
from tkinter import filedialog
from typing import TYPE_CHECKING, List, Optional

import customtkinter as ctk

from app.constants import UserRole
from app.theme import (
    COLORS,
    SPACING,
    TYPOGRAPHY,
    configure_button_style,
    configure_card_style,
    configure_input_style,
)
from dialogs.confirm_dialog import ConfirmDialog, InfoDialog
from models.user import User
from services.user_service import UserService
from views.base_view import BaseView

if TYPE_CHECKING:
    from app.application import PolicyHubApp

# Try to import tksheet
try:
    from tksheet import Sheet
    TKSHEET_AVAILABLE = True
except ImportError:
    TKSHEET_AVAILABLE = False

# Try to import openpyxl for export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class UsersSettingsView(BaseView):
    """
    User management view for administrators.

    Features:
    - View all users in a table
    - Add new users
    - Edit existing users
    - Activate/deactivate users
    - Reset passwords
    """

    def __init__(self, parent: ctk.CTkFrame, app: "PolicyHubApp"):
        """
        Initialize the users settings view.

        Args:
            parent: Parent widget
            app: Main application instance
        """
        super().__init__(parent, app)
        self.user_service = UserService(app.db)
        self.users: List[User] = []
        self.filtered_users: List[User] = []
        self.show_inactive = False
        self.search_term = ""
        self._search_after_id = None  # For debouncing search
        self._build_ui()

    def _build_ui(self) -> None:
        """Build the users management UI."""
        # Header
        header = ctk.CTkFrame(self, fg_color="transparent")
        header.pack(fill="x", padx=SPACING.WINDOW_PADDING, pady=(SPACING.WINDOW_PADDING, 10))

        title = ctk.CTkLabel(
            header,
            text="User Management",
            font=TYPOGRAPHY.section_heading,
            text_color=COLORS.TEXT_PRIMARY,
        )
        title.pack(side="left")

        # Add button
        add_btn = ctk.CTkButton(
            header,
            text="+ Add User",
            command=self._on_add_user,
            width=120,
        )
        configure_button_style(add_btn, "primary")
        add_btn.pack(side="right")

        # Toolbar row (search, export, import buttons)
        toolbar = ctk.CTkFrame(self, fg_color="transparent")
        toolbar.pack(fill="x", padx=SPACING.WINDOW_PADDING, pady=(0, 10))

        # Search box
        search_frame = ctk.CTkFrame(toolbar, fg_color="transparent")
        search_frame.pack(side="left")

        self.search_entry = ctk.CTkEntry(
            search_frame,
            width=250,
            height=32,
            placeholder_text="Search users...",
            font=TYPOGRAPHY.body,
        )
        configure_input_style(self.search_entry)
        self.search_entry.pack(side="left")
        self.search_entry.bind("<KeyRelease>", self._on_search_changed)

        # Show inactive toggle
        self.inactive_var = ctk.BooleanVar(value=False)
        inactive_check = ctk.CTkCheckBox(
            toolbar,
            text="Show inactive",
            variable=self.inactive_var,
            command=self._refresh_table,
            font=TYPOGRAPHY.small,
            text_color=COLORS.TEXT_SECONDARY,
        )
        inactive_check.pack(side="left", padx=(20, 0))

        # Bulk deactivate button (initially hidden)
        self.bulk_deactivate_btn = ctk.CTkButton(
            toolbar,
            text="Deactivate Selected",
            command=self._on_bulk_deactivate,
            width=140,
            state="disabled",
        )
        configure_button_style(self.bulk_deactivate_btn, "secondary")
        self.bulk_deactivate_btn.pack(side="right", padx=(8, 0))

        # Import button
        import_btn = ctk.CTkButton(
            toolbar,
            text="Import",
            command=self._on_import_users,
            width=80,
        )
        configure_button_style(import_btn, "secondary")
        import_btn.pack(side="right", padx=(8, 0))

        # Template button
        template_btn = ctk.CTkButton(
            toolbar,
            text="Template",
            command=self._on_download_template,
            width=80,
        )
        configure_button_style(template_btn, "secondary")
        template_btn.pack(side="right", padx=(8, 0))

        # Export button
        export_btn = ctk.CTkButton(
            toolbar,
            text="Export",
            command=self._on_export_users,
            width=80,
        )
        configure_button_style(export_btn, "secondary")
        export_btn.pack(side="right")

        # Authentication Settings card
        self._build_auth_settings_card()

        # Table container
        table_card = ctk.CTkFrame(self, fg_color=COLORS.CARD)
        configure_card_style(table_card, with_shadow=True)
        table_card.pack(
            fill="both",
            expand=True,
            padx=SPACING.WINDOW_PADDING,
            pady=(0, SPACING.WINDOW_PADDING),
        )

        # Build table
        if TKSHEET_AVAILABLE:
            self._build_tksheet_table(table_card)
        else:
            self._build_fallback_table(table_card)

        # Status bar
        self.status_label = ctk.CTkLabel(
            table_card,
            text="Loading...",
            font=TYPOGRAPHY.small,
            text_color=COLORS.TEXT_SECONDARY,
        )
        self.status_label.pack(side="bottom", anchor="w", padx=SPACING.CARD_PADDING, pady=8)

    def _build_auth_settings_card(self) -> None:
        """Build the authentication settings card."""
        from services.settings_service import SettingsService

        # Initialize settings service
        self.settings_service = SettingsService(self.app.db)

        # Auth settings card
        auth_card = ctk.CTkFrame(self, fg_color=COLORS.CARD)
        configure_card_style(auth_card, with_shadow=True)
        auth_card.pack(
            fill="x",
            padx=SPACING.WINDOW_PADDING,
            pady=(0, 12),
        )

        # Card content
        content = ctk.CTkFrame(auth_card, fg_color="transparent")
        content.pack(fill="x", padx=SPACING.CARD_PADDING, pady=SPACING.CARD_PADDING)

        # Card title
        title = ctk.CTkLabel(
            content,
            text="Authentication Settings",
            font=TYPOGRAPHY.section_heading,
            text_color=COLORS.TEXT_PRIMARY,
        )
        title.pack(anchor="w", pady=(0, 12))

        # Settings row container
        settings_row = ctk.CTkFrame(content, fg_color="transparent")
        settings_row.pack(fill="x")

        # Left side: Require Login toggle
        left_frame = ctk.CTkFrame(settings_row, fg_color="transparent")
        left_frame.pack(side="left", fill="x", expand=True)

        # Require Login setting
        login_frame = ctk.CTkFrame(left_frame, fg_color="transparent")
        login_frame.pack(anchor="w")

        self.require_login_var = ctk.BooleanVar(
            value=self.settings_service.get_require_login()
        )
        require_login_switch = ctk.CTkSwitch(
            login_frame,
            text="Require Login",
            variable=self.require_login_var,
            command=self._on_require_login_changed,
            font=TYPOGRAPHY.body,
            text_color=COLORS.TEXT_PRIMARY,
        )
        require_login_switch.pack(side="left")

        # Help text for require login
        login_help = ctk.CTkLabel(
            left_frame,
            text="When disabled, the app will automatically log in as the admin user.",
            font=TYPOGRAPHY.small,
            text_color=COLORS.TEXT_MUTED,
        )
        login_help.pack(anchor="w", pady=(4, 0))

        # Right side: Master Password button
        right_frame = ctk.CTkFrame(settings_row, fg_color="transparent")
        right_frame.pack(side="right")

        master_pwd_btn = ctk.CTkButton(
            right_frame,
            text="Change Master Password",
            command=self._on_change_master_password,
            width=180,
        )
        configure_button_style(master_pwd_btn, "secondary")
        master_pwd_btn.pack()

        # Help text for master password
        pwd_help = ctk.CTkLabel(
            right_frame,
            text="Safety net for forgotten admin passwords",
            font=TYPOGRAPHY.small,
            text_color=COLORS.TEXT_MUTED,
        )
        pwd_help.pack(anchor="e", pady=(4, 0))

    def _on_require_login_changed(self) -> None:
        """Handle require login toggle change."""
        require_login = self.require_login_var.get()

        try:
            self.settings_service.set_require_login(require_login)

            status = "enabled" if require_login else "disabled"
            InfoDialog.show_success(
                self.winfo_toplevel(),
                "Setting Updated",
                f"Login requirement has been {status}.\n\n"
                f"{'Users will need to log in to access the application.' if require_login else 'The application will auto-login as the admin user.'}",
            )

        except PermissionError:
            # Revert the switch
            self.require_login_var.set(not require_login)
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Permission Denied",
                "You do not have permission to change this setting.",
            )
        except Exception as e:
            # Revert the switch
            self.require_login_var.set(not require_login)
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Error",
                f"Failed to update setting: {str(e)}",
            )

    def _on_change_master_password(self) -> None:
        """Handle change master password button click."""
        from dialogs.master_password_dialog import MasterPasswordDialog

        dialog = MasterPasswordDialog(self.winfo_toplevel(), self.app.db)
        dialog.show()

    def _build_tksheet_table(self, parent) -> None:
        """Build the tksheet table."""
        self.table_card = parent  # Store reference for resize calculations

        self.table = Sheet(
            parent,
            headers=["Username", "Full Name", "Email", "Role", "Status", "Last Login"],
            header_bg=COLORS.MUTED,
            header_fg=COLORS.TEXT_PRIMARY,
            header_font=("Segoe UI", 11, "bold"),
            header_grid_fg=COLORS.BORDER,
            table_bg=COLORS.CARD,
            table_fg=COLORS.TEXT_PRIMARY,
            font=("Segoe UI", 11, "normal"),
            table_grid_fg=COLORS.BORDER,
            table_selected_cells_bg=COLORS.PRIMARY,
            table_selected_cells_fg=COLORS.PRIMARY_FOREGROUND,
            index_bg=COLORS.MUTED,
            top_left_bg=COLORS.MUTED,
            outline_thickness=1,
            outline_color=COLORS.BORDER,
            frame_bg=COLORS.CARD,
            show_row_index=False,
            show_top_left=False,
            default_row_height=36,
        )
        self.table.pack(fill="both", expand=True, padx=SPACING.CARD_PADDING, pady=(SPACING.CARD_PADDING, 0))

        # Enable features (multi-select for bulk operations)
        self.table.enable_bindings(
            "row_select",
            "column_width_resize",
            "arrowkeys",
            "ctrl_select",
            "shift_select",
        )

        # Enable text wrapping for better display of long text
        self.table.set_options(
            auto_resize_row_height=True,  # Auto-resize rows for wrapped text
        )

        # Track selection changes for bulk deactivate button
        self.table.extra_bindings("cell_select", self._on_selection_changed)
        self.table.extra_bindings("shift_cell_select", self._on_selection_changed)
        self.table.extra_bindings("ctrl_cell_select", self._on_selection_changed)
        self.table.extra_bindings("deselect", self._on_selection_changed)

        # Column proportions (weights) - Email gets more space
        # [Username, Full Name, Email, Role, Status, Last Login]
        self._column_weights = [1.0, 1.5, 2.0, 0.8, 0.7, 1.2]
        self._column_min_widths = [100, 140, 160, 80, 70, 120]

        # Set initial column widths
        self._resize_columns()

        # Bind to configure event for responsive resizing
        parent.bind("<Configure>", self._on_container_resize)
        self._resize_scheduled = False

        # Double-click to edit
        self.table.bind("<Double-Button-1>", self._on_double_click)

        # Right-click context menu
        self.table.bind("<Button-3>", self._show_context_menu)

    def _on_container_resize(self, event=None) -> None:
        """Handle container resize - debounced."""
        if not hasattr(self, '_resize_scheduled'):
            self._resize_scheduled = False
        if not self._resize_scheduled:
            self._resize_scheduled = True
            self.after(100, self._do_resize)

    def _do_resize(self) -> None:
        """Perform the actual resize."""
        self._resize_scheduled = False
        self._resize_columns()

    def _resize_columns(self) -> None:
        """Resize columns to fill available width proportionally."""
        if not TKSHEET_AVAILABLE or not self.table:
            return

        try:
            # Get available width (container width minus padding and scrollbar)
            available_width = self.table_card.winfo_width() - (SPACING.CARD_PADDING * 2) - 20

            if available_width < 400:  # Not yet rendered or too small
                # Use fallback widths
                fallback_widths = [120, 180, 200, 100, 80, 150]
                for i, width in enumerate(fallback_widths):
                    self.table.column_width(i, width)
                return

            # Calculate total weight
            total_weight = sum(self._column_weights)

            # Calculate widths based on weights
            for i, (weight, min_width) in enumerate(zip(self._column_weights, self._column_min_widths)):
                proportional_width = int((weight / total_weight) * available_width)
                final_width = max(proportional_width, min_width)
                self.table.column_width(i, final_width)

        except Exception:
            pass  # Ignore errors during resize

    def _build_fallback_table(self, parent) -> None:
        """Build a fallback table without tksheet."""
        self.table_scroll = ctk.CTkScrollableFrame(parent, fg_color="transparent")
        self.table_scroll.pack(fill="both", expand=True, padx=SPACING.CARD_PADDING, pady=SPACING.CARD_PADDING)

        # Header row
        header = ctk.CTkFrame(self.table_scroll, fg_color=COLORS.MUTED)
        header.pack(fill="x")

        headers = ["Username", "Full Name", "Role", "Status"]
        for text in headers:
            ctk.CTkLabel(
                header,
                text=text,
                font=TYPOGRAPHY.get_font(11, "bold"),
                width=150,
                anchor="w",
            ).pack(side="left", padx=10, pady=8)

        self.table = None

    def _update_fallback_table(self) -> None:
        """Update the fallback table."""
        # Clear existing rows
        children = self.table_scroll.winfo_children()
        for child in children[1:]:
            child.destroy()

        for user in self.users:
            row = ctk.CTkFrame(self.table_scroll, fg_color="transparent", cursor="hand2")
            row.pack(fill="x")
            row.bind("<Double-Button-1>", lambda e, u=user: self._edit_user(u))

            values = [
                user.username,
                user.full_name,
                user.role_display,
                "Active" if user.is_active else "Inactive",
            ]

            for val in values:
                label = ctk.CTkLabel(
                    row,
                    text=val,
                    font=TYPOGRAPHY.body,
                    width=150,
                    anchor="w",
                )
                label.pack(side="left", padx=10, pady=6)
                label.bind("<Double-Button-1>", lambda e, u=user: self._edit_user(u))

            divider = ctk.CTkFrame(self.table_scroll, fg_color=COLORS.BORDER, height=1)
            divider.pack(fill="x")

    def _refresh_table(self) -> None:
        """Reload user data."""
        self.show_inactive = self.inactive_var.get()
        self.users = self.user_service.get_all_users(include_inactive=self.show_inactive)
        self._apply_filter()

    def _apply_filter(self) -> None:
        """Apply search filter to users and update table."""
        search = self.search_term.lower().strip()

        if search:
            self.filtered_users = [
                u for u in self.users
                if search in u.username.lower()
                or search in u.full_name.lower()
                or (u.email and search in u.email.lower())
            ]
        else:
            self.filtered_users = self.users.copy()

        if TKSHEET_AVAILABLE and self.table:
            data = []
            for user in self.filtered_users:
                from utils.dates import format_datetime
                last_login = format_datetime(user.last_login) if user.last_login else "Never"
                data.append([
                    user.username,
                    user.full_name,
                    user.email or "",
                    user.role_display,
                    "Active" if user.is_active else "Inactive",
                    last_login,
                ])
            self.table.set_sheet_data(data)
            self.table.deselect()  # Clear selection after refresh
        else:
            self._update_fallback_table()

        # Update status and bulk button
        self.status_label.configure(text=f"Showing {len(self.filtered_users)} user(s)")
        self.bulk_deactivate_btn.configure(state="disabled")

    def _on_search_changed(self, event=None) -> None:
        """Handle search input changes with debouncing."""
        # Cancel previous scheduled search
        if self._search_after_id:
            self.after_cancel(self._search_after_id)

        # Schedule new search after 300ms
        self._search_after_id = self.after(300, self._perform_search)

    def _perform_search(self) -> None:
        """Perform the actual search."""
        self.search_term = self.search_entry.get()
        self._apply_filter()

    def _on_selection_changed(self, event=None) -> None:
        """Handle table selection changes."""
        if not TKSHEET_AVAILABLE or not self.table:
            return

        selected_rows = self.table.get_selected_rows()
        if selected_rows:
            self.bulk_deactivate_btn.configure(state="normal")
        else:
            self.bulk_deactivate_btn.configure(state="disabled")

    def _on_add_user(self) -> None:
        """Handle add user button click."""
        from dialogs.user_dialog import UserDialog

        dialog = UserDialog(self.winfo_toplevel(), self.app.db)
        result = dialog.show()
        if result:
            self._refresh_table()

    def _on_double_click(self, event) -> None:
        """Handle double-click on table row."""
        if not TKSHEET_AVAILABLE or not self.table:
            return

        selected = self.table.get_currently_selected()
        if selected and len(self.filtered_users) > 0:
            row = selected.row
            if 0 <= row < len(self.filtered_users):
                self._edit_user(self.filtered_users[row])

    def _edit_user(self, user: User) -> None:
        """Open edit dialog for a user."""
        from dialogs.user_dialog import UserDialog

        dialog = UserDialog(self.winfo_toplevel(), self.app.db, user=user)
        result = dialog.show()
        if result:
            self._refresh_table()

    def _show_context_menu(self, event) -> None:
        """Show context menu on right-click."""
        if not TKSHEET_AVAILABLE or not self.table:
            return

        # Get clicked row
        try:
            row = self.table.identify_row(event, allow_end=False)
            if row is None or row >= len(self.filtered_users):
                return
            user = self.filtered_users[row]
        except Exception:
            return

        # Create context menu
        menu = ctk.CTkFrame(
            self.winfo_toplevel(),
            fg_color=COLORS.CARD,
            border_width=1,
            border_color=COLORS.BORDER,
            corner_radius=SPACING.CORNER_RADIUS,
        )

        # Menu items
        edit_btn = ctk.CTkButton(
            menu,
            text="Edit User",
            command=lambda: [menu.destroy(), self._edit_user(user)],
            fg_color="transparent",
            text_color=COLORS.TEXT_PRIMARY,
            hover_color=COLORS.MUTED,
            anchor="w",
            height=32,
        )
        edit_btn.pack(fill="x", padx=4, pady=2)

        if user.is_active:
            deactivate_btn = ctk.CTkButton(
                menu,
                text="Deactivate",
                command=lambda: [menu.destroy(), self._toggle_active(user, False)],
                fg_color="transparent",
                text_color=COLORS.TEXT_PRIMARY,
                hover_color=COLORS.MUTED,
                anchor="w",
                height=32,
            )
            deactivate_btn.pack(fill="x", padx=4, pady=2)
        else:
            activate_btn = ctk.CTkButton(
                menu,
                text="Activate",
                command=lambda: [menu.destroy(), self._toggle_active(user, True)],
                fg_color="transparent",
                text_color=COLORS.TEXT_PRIMARY,
                hover_color=COLORS.MUTED,
                anchor="w",
                height=32,
            )
            activate_btn.pack(fill="x", padx=4, pady=2)

        reset_btn = ctk.CTkButton(
            menu,
            text="Reset Password",
            command=lambda: [menu.destroy(), self._reset_password(user)],
            fg_color="transparent",
            text_color=COLORS.TEXT_PRIMARY,
            hover_color=COLORS.MUTED,
            anchor="w",
            height=32,
        )
        reset_btn.pack(fill="x", padx=4, pady=2)

        # Position menu
        menu.place(x=event.x_root - self.winfo_toplevel().winfo_x(),
                   y=event.y_root - self.winfo_toplevel().winfo_y())

        # Close menu on click outside
        def close_menu(e):
            if menu.winfo_exists():
                menu.destroy()

        self.winfo_toplevel().bind("<Button-1>", close_menu, add="+")

    def _toggle_active(self, user: User, activate: bool) -> None:
        """Activate or deactivate a user."""
        try:
            if activate:
                self.user_service.activate_user(user.user_id)
                InfoDialog.show_success(
                    self.winfo_toplevel(),
                    "User Activated",
                    f"User '{user.username}' has been activated.",
                )
            else:
                # Check if trying to deactivate self
                from core.session import SessionManager
                session = SessionManager.get_instance()
                if session.current_user and session.current_user.user_id == user.user_id:
                    InfoDialog.show_error(
                        self.winfo_toplevel(),
                        "Cannot Deactivate",
                        "You cannot deactivate your own account.",
                    )
                    return

                self.user_service.deactivate_user(user.user_id)
                InfoDialog.show_success(
                    self.winfo_toplevel(),
                    "User Deactivated",
                    f"User '{user.username}' has been deactivated.",
                )
            self._refresh_table()
        except Exception as e:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Error",
                str(e),
            )

    def _reset_password(self, user: User) -> None:
        """Reset a user's password."""
        from dialogs.password_reset_dialog import PasswordResetDialog

        dialog = PasswordResetDialog(self.winfo_toplevel(), self.app.db, user=user)
        result = dialog.show()
        if result:
            InfoDialog.show_success(
                self.winfo_toplevel(),
                "Password Reset",
                f"Password for '{user.username}' has been reset successfully.",
            )

    def on_show(self) -> None:
        """Called when the view becomes visible."""
        super().on_show()
        self._refresh_table()
        # Force resize after view is fully rendered to ensure columns fill space
        self.after(50, self._resize_columns)

    def _on_export_users(self) -> None:
        """Export user list to Excel."""
        if not OPENPYXL_AVAILABLE:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Export Unavailable",
                "Excel export requires openpyxl library.",
            )
            return

        # Get export data
        export_data = self.user_service.get_users_for_export(
            include_inactive=self.inactive_var.get()
        )

        if not export_data:
            InfoDialog.show_info(
                self.winfo_toplevel(),
                "No Data",
                "No users to export.",
            )
            return

        # File dialog
        file_path = filedialog.asksaveasfilename(
            parent=self.winfo_toplevel(),
            title="Export Users",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
            initialfile="users_export.xlsx",
        )

        if not file_path:
            return

        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Users"

            # Styles
            header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style="thin", color="E5E7EB"),
                right=Side(style="thin", color="E5E7EB"),
                top=Side(style="thin", color="E5E7EB"),
                bottom=Side(style="thin", color="E5E7EB"),
            )

            # Headers
            headers = ["Username", "Full Name", "Email", "Role", "Status", "Created", "Last Login"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            # Data
            for row_idx, user_data in enumerate(export_data, 2):
                ws.cell(row=row_idx, column=1, value=user_data["username"]).border = thin_border
                ws.cell(row=row_idx, column=2, value=user_data["full_name"]).border = thin_border
                ws.cell(row=row_idx, column=3, value=user_data["email"]).border = thin_border
                ws.cell(row=row_idx, column=4, value=user_data["role"]).border = thin_border
                ws.cell(row=row_idx, column=5, value=user_data["status"]).border = thin_border
                ws.cell(row=row_idx, column=6, value=user_data["created_at"]).border = thin_border
                ws.cell(row=row_idx, column=7, value=user_data["last_login"]).border = thin_border

            # Column widths
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 30
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 12
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 15

            # Save
            wb.save(file_path)

            InfoDialog.show_success(
                self.winfo_toplevel(),
                "Export Complete",
                f"Exported {len(export_data)} users to:\n{file_path}",
            )

        except Exception as e:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Export Error",
                f"Failed to export users: {str(e)}",
            )

    def _on_download_template(self) -> None:
        """Download CSV import template."""
        file_path = filedialog.asksaveasfilename(
            parent=self.winfo_toplevel(),
            title="Save Import Template",
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
            initialfile="user_import_template.csv",
        )

        if not file_path:
            return

        try:
            template = self.user_service.get_csv_template()
            Path(file_path).write_text(template, encoding="utf-8")

            InfoDialog.show_success(
                self.winfo_toplevel(),
                "Template Saved",
                f"Import template saved to:\n{file_path}\n\nEdit the file to add users, then use Import to add them.",
            )

        except Exception as e:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Error",
                f"Failed to save template: {str(e)}",
            )

    def _on_import_users(self) -> None:
        """Import users from CSV file."""
        file_path = filedialog.askopenfilename(
            parent=self.winfo_toplevel(),
            title="Import Users from CSV",
            filetypes=[("CSV Files", "*.csv"), ("All Files", "*.*")],
        )

        if not file_path:
            return

        # Validate first
        result = self.user_service.validate_import_data(Path(file_path))

        if not result.success:
            # Show all errors
            error_text = "\n".join(result.error_messages[:20])  # Limit to 20 errors
            if len(result.errors) > 20:
                error_text += f"\n... and {len(result.errors) - 20} more errors"

            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Import Validation Failed",
                f"The following errors were found:\n\n{error_text}",
            )
            return

        # Confirm import
        confirm = ConfirmDialog.show(
            self.winfo_toplevel(),
            "Confirm Import",
            f"Ready to import {result.imported_count} user(s).\n\n"
            "Users will be required to change their password on first login.\n\n"
            "Continue with import?",
        )

        if not confirm:
            return

        # Perform import
        try:
            import_result = self.user_service.import_users_from_csv(Path(file_path))

            if import_result.success:
                InfoDialog.show_success(
                    self.winfo_toplevel(),
                    "Import Complete",
                    f"Successfully imported {import_result.imported_count} user(s).",
                )
                self._refresh_table()
            else:
                error_text = "\n".join(import_result.error_messages)
                InfoDialog.show_error(
                    self.winfo_toplevel(),
                    "Import Failed",
                    f"Import failed:\n\n{error_text}",
                )

        except PermissionError:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Permission Denied",
                "You do not have permission to import users.",
            )
        except Exception as e:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Import Error",
                f"Failed to import users: {str(e)}",
            )

    def _on_bulk_deactivate(self) -> None:
        """Deactivate selected users."""
        if not TKSHEET_AVAILABLE or not self.table:
            return

        selected_rows = self.table.get_selected_rows()
        if not selected_rows:
            return

        # Get selected users
        selected_users = [
            self.filtered_users[row]
            for row in selected_rows
            if 0 <= row < len(self.filtered_users)
        ]

        if not selected_users:
            return

        # Confirm
        usernames = ", ".join(u.username for u in selected_users[:5])
        if len(selected_users) > 5:
            usernames += f" and {len(selected_users) - 5} more"

        confirm = ConfirmDialog.show(
            self.winfo_toplevel(),
            "Confirm Bulk Deactivation",
            f"Are you sure you want to deactivate {len(selected_users)} user(s)?\n\n"
            f"Users: {usernames}",
        )

        if not confirm:
            return

        # Perform bulk deactivation
        user_ids = [u.user_id for u in selected_users]

        try:
            count, errors = self.user_service.bulk_deactivate_users(user_ids)

            if errors:
                error_text = "\n".join(errors[:10])
                if len(errors) > 10:
                    error_text += f"\n... and {len(errors) - 10} more"

                InfoDialog.show_info(
                    self.winfo_toplevel(),
                    "Partial Success",
                    f"Deactivated {count} user(s).\n\nSome users could not be deactivated:\n{error_text}",
                )
            else:
                InfoDialog.show_success(
                    self.winfo_toplevel(),
                    "Users Deactivated",
                    f"Successfully deactivated {count} user(s).",
                )

            self._refresh_table()

        except PermissionError:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Permission Denied",
                "You do not have permission to deactivate users.",
            )
        except Exception as e:
            InfoDialog.show_error(
                self.winfo_toplevel(),
                "Error",
                f"Failed to deactivate users: {str(e)}",
            )
